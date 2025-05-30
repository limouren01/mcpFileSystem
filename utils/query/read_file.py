import os
import csv
import json
from docx import Document
from openpyxl import load_workbook
from config.config import is_path_allowed, ALLOWED_READ_DIR

def read_file(file_path: str, allowed_read_dir: str) -> dict:
    """读取文件（严格路径校验），支持多种常见文件格式"""
    abs_path = os.path.abspath(file_path)
    if not is_path_allowed(abs_path, allowed_read_dir):
        return {"success": False, "message": "文件路径越权访问"}

    if not os.path.isfile(abs_path):
        return {"success": False, "message": "指定的文件不存在"}

    try:
        if file_path.endswith('.csv'):
            with open(abs_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                content = [row for row in reader]
        elif file_path.endswith('.docx'):
            doc = Document(abs_path)
            content = '\n'.join([para.text for para in doc.paragraphs])
        elif file_path.endswith('.xlsx'):
            wb = load_workbook(filename=abs_path, data_only=True)
            content = {}
            for sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    rows.append(row)
                content[sheetname] = rows
        elif file_path.endswith('.json'):
            with open(abs_path, 'r', encoding='utf-8') as f:
                content = json.load(f)  # 使用 json 模块解析 JSON 文件
        else:
            with open(abs_path, 'r', encoding='utf-8') as f:
                content = f.read()
        return {"success": True, "content": content}
    except Exception as e:
        return {"success": False, "message": f"文件读取失败: {str(e)}"}